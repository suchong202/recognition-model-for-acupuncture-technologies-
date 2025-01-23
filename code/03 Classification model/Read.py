import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import random
import cv2
import mediapipe as mp
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
import math

#Randomly select
def lottery_draw(n, start, end):
    numbers = list(range(start, end + 1))
    lucky_numbers = random.sample(numbers, n)
    return lucky_numbers

#Turn execl to list
def readexecl(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    D = []
    n = sheet.max_column
    for i in range(0,n):
        A=[]
        D.append(A)

    for row in sheet.iter_rows():

        row_data = [cell.value for cell in row]

        for i in range(0, n):

            #D[i].append(float(row_data[i] - row_data[i % 3]))
            D[i].append(float(row_data[i]))
    return D

def extend(r, n):
    num1 = int(n/len(r))
    num2 = int(n%len(r))

    L =[]

    if n>=len(r):
        for i in range(0, num1):
            for j in range(0, len(r)):
                L.append(r[j])

        for j in range(0, num2):
            L.append(r[j])

    else:
        for j in range(0, num2):
            L.append(r[j])

    return L

def cut(r, n):

   a = int((len(r) / 2) - (n*len(r))/2)
   b = int((len(r) / 2) + (n*len(r))/2)
   L = r[a:b]

   return L

#normalization
def guiyi(r):
    min_val = min(r)
    max_val = max(r)
    normalized_arr = [(x - min_val) / (max_val - min_val) for x in r]

    return normalized_arr


mpHands = mp.solutions.hands
hands = mpHands.Hands(static_image_mode=False,
                      max_num_hands=1,
                      min_detection_confidence=0.75,
                      min_tracking_confidence=0.5
                      )
mpDraw = mp.solutions.drawing_utils

def getxyz(datapath,savepath):

        wb = openpyxl.Workbook()  # 创建一个excel文件
        sheet = wb.active  # 获得一个的工作表

        cap = cv2.VideoCapture(datapath)
        success, frame = cap.read()
        if not success:
            print("无法读取视频")
        else:
            # 获取图像宽度和高度
            width = frame.shape[1]
            height = frame.shape[0]


        #统计帧数
        i = 0

        while True:

            sussess, img = cap.read()

            if not sussess:
                break

            imgRGB = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            results = hands.process(imgRGB)


            if results.multi_hand_landmarks:

                for hand_landmarks in results.multi_hand_landmarks:
                    print('hand_landmarks:',hand_landmarks)

                    print("\n")
                    m = hand_landmarks.landmark

                    A = []

                    x0 = 0
                    y0 = 0
                    z0 = 0
                    x1 = 0
                    y1 = 0
                    z1 = 0

                    for id, lm in enumerate(m):

                        print("n={}, x = {}, y = {}, z = {}".format(id, lm.x, lm.y, lm.z))
                        if id == 0:
                            x0 = lm.x
                            y0 = lm.y
                            z0 = lm.z
                        if id == 1:
                            x1 = lm.x
                            y1 = lm.y
                            z1 = lm.z
                        if id<=8 and id>=1:
                            k=math.sqrt((x1 - x0) * (x1 - x0) + (y1 - y0) * (y1 - y0) + (z1 - z0) * (z1 - z0))
                            A.append((lm.x-x0)/k)
                            A.append((lm.y-y0)/k)
                            A.append((lm.z-z0)/k)
                    sheet.append(A)

                    # 创建一个500*500,3颜色通道图片的numpy矩阵
                    img = np.zeros((height,width,  3), dtype=np.uint8)
                    img[:] = (255, 255, 255)  # 白色背景

                    # 关键点可视化
                    handLmsStyle = mpDraw.DrawingSpec(color=(0, 0, 255), thickness=3)
                    handConStyle = mpDraw.DrawingSpec(color=(0, 255, 0), thickness=2)
                    mpDraw.draw_landmarks(img, hand_landmarks, mpHands.HAND_CONNECTIONS, handLmsStyle, handConStyle)

            #cv2.imshow('image', img)

            img = np.zeros((width,height, 3), np.uint8)
            img.fill(0)
            i = i + 1
            if cv2.waitKey(1) & 0xFF == 27:
                break
            print("共{}帧,第{}帧".format(cap.get(7), i))

        cap.release()

        wb.save(savepath)

# turn video to execl
def videotoexecl(url1,url2):
    file0 = os.listdir(url1)
    for f0 in file0:
        # 字符串拼接
        test_url1 = url1 + f0
        test_url2 = url2 + f0

        file1 = os.listdir(test_url1)
        for f1 in file1:
            datapath = test_url1 + '/' + f1
            savepath = test_url2 + '/' + f1

            print(datapath,savepath[0:-3]+'xlsx')
            getxyz(datapath, savepath[0:-3]+'xlsx')



if __name__ == '__main__':

    datapath = './1.mp4'
    savepath = './1.xlsx'
    #getxyz(datapath,savepath)

    test1 = './video/Expert/'
    test2 = './data/V/Expert/'
    videotoexecl(test1,test2)
