import openpyxl
from openpyxl import load_workbook
import numpy as np
import tensorflow as tf
from sklearn.model_selection import train_test_split
from keras import Sequential
from keras.layers import Dense
from keras.layers import Dropout

#Univariate feature selection
def compare1(x,y,feature_names):
    from sklearn.feature_selection import SelectKBest, f_regression
    from sklearn.pipeline import make_pipeline
    selector = SelectKBest(f_regression, k=2)
    pipeline = make_pipeline(selector)
    pipeline.fit(x, y)
    selected_features = [feature_names[i] for i in range(len(feature_names)) if selector.get_support()[i]]
    print("Selected features1:", selected_features)

    P=[]
    for i in range (0,len(feature_names)):
        if feature_names[i] in selected_features:
            P.append(1)
        else:
            P.append(0)
    print(P)

    return P

#Feature Importance from Tree-Based Models
def compare2(x,y,feature_names):
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.feature_selection import SelectFromModel

    estimator = RandomForestClassifier(n_estimators=100, random_state=42)
    estimator.fit(x, y)
    sfm = SelectFromModel(estimator=estimator, threshold='median', prefit=True, max_features=2)
    X_transformed = sfm.transform(x)
    print("Selected features2:", sfm.get_support())

    P = []
    for i in range(0, len(feature_names)):
        if sfm.get_support()[i] ==True:
            P.append(1)
        else:
            P.append(0)
    print(P)

    return P

# Recursive Feature Elimination
def compare3(x,y,feature_names):
    from sklearn.datasets import make_friedman1
    from sklearn.feature_selection import RFE
    from sklearn.svm import SVR
    estimator = SVR(kernel="linear")
    selector = RFE(estimator, n_features_to_select=5, step=1, verbose=0)
    selector.fit(x,y)
    print("Selected features:", selector.support_)

    P = []
    for i in range(0, len(feature_names)):
        if selector.support_[i] == True:
            P.append(1)
        else:
            P.append(0)
    print(P)

    return P

# L1-based feature selection
def compare4(x,y,feature_names):
    from sklearn.linear_model import Lasso
    from sklearn.preprocessing import StandardScaler
    scaler = StandardScaler()

    lasso = Lasso(alpha=0.2)
    lasso.fit(x, y)
    lasso.coef_
    print("Selected features4:", lasso.coef_)

    P = lasso.coef_
    return P

# L2-based feature selection
def compare5(x,y,feature_names):

    from sklearn.preprocessing import StandardScaler
    from sklearn.linear_model import LinearRegression
    scaler = StandardScaler()
    lr = LinearRegression()
    lr.fit(x, y)
    print("Selected features5:", lr.coef_)
    P = lr.coef_
    return P

# Mutual Information
def compare6(x,y,feature_names):
    from sklearn.feature_selection import mutual_info_classif
    n_features = x.shape[1]
    mi = mutual_info_classif(x, y)
    threshold = 0.1
    selected_features = [i for i, m in enumerate(mi) if m > threshold]
    selected_x = x[:, selected_features]
    print("Selected features6:", selected_features)
    P = []
    for i in range(0, len(feature_names)):
        if (int(i)-1) in selected_features:
            P.append(1)
        else:
            P.append(0)
    print(P)
    return P

# Variance Threshold
def compare7(x,y,feature_names):
    from sklearn.feature_selection import VarianceThreshold

    features = np.array(x)
    target = np.array(y)

    sel = VarianceThreshold(threshold=0.0)
    sel.fit_transform(features)

    print("Selected features7:", sel.get_support())

    P = []
    for i in range(0, len(feature_names)):
        if sel.get_support()[i] == True:
            P.append(1)
        else:
            P.append(0)
    print(P)

    return P

# Principal Component Analysis
def compare8(x,y,feature_names):
    from sklearn.decomposition import PCA

    pca = PCA(n_components=2)
    x_reduced = pca.fit_transform(x)

    print("Selected features8:", x_reduced)

    return x_reduced

# DNN
def net(path):

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values = []
    y_values = []
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data) - 1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]
    x_values = np.array(x_values[1:])
    y_values = np.array(y_values[1:])

    print(features)
    P=compare4(x_values,y_values,features)
    x_values = np.array(bian(x_values, P))

    x_train, x_test, y_train, y_test = train_test_split(x_values, y_values, test_size=0.2, random_state=42,shuffle=True)

    model_name1 = 'relu'
    model_name2 = 'sigmoid'
    model_name3 = 'tanh'
    model_name4 = 'softmax'
    model_name5 = 'elu'
    model_name6 = 'swish'
    model_name = model_name2
    model = Sequential()
    model.add(Dense(8, input_shape=(len(features),)))
    model.add(Dense(16, activation=model_name))
    model.add(Dense(32, activation=model_name))
    model.add(Dropout(0.2))
    model.add(Dense(32, activation=model_name))
    model.add(Dense(16, activation=model_name))
    model.add(Dense(8, activation=model_name))
    model.add(Dense(1, activation=model_name))
    model.summary()

    model.compile(loss=tf.keras.losses.logcosh, metrics=['accuracy'], optimizer=tf.optimizers.Adam(lr=1e-03))
    history = model.fit(x_train, y_train, batch_size=16, epochs=200, validation_data=(x_test, y_test),validation_freq=1)
    model.save('./output/model.h5')

def bian(X,P):

    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x

if __name__ == '__main__':

    path = './data/heart.xlsx'
    net(path)