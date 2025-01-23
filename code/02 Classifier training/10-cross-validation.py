import matplotlib
matplotlib.use('TkAgg')
import numpy as np
import openpyxl
import tensorflow as tf
from keras import Sequential
from keras.layers import Dense
from keras.layers import Dropout
from openpyxl import load_workbook
from sklearn.model_selection import KFold

#10-Cross-validation
def net(P:[],path,maxaccuracy):

    model_name1 = 'relu'
    model_name2 = 'sigmoid'
    model_name3 = 'tanh'
    model_name4 = 'softmax'
    model_name5 = 'elu'
    model_name6 = 'swish'

    model_name = model_name2

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values=[]
    y_values=[]
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data)-1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]
    x_values = np.array(x_values[1:])
    x_values = np.array(bian(x_values,P))
    y_values = np.array(y_values[1:])

    #折数
    w =  10
    kf = KFold(n_splits=w, random_state=2000, shuffle=True)

    curr_score = 0

    for train_index, test_index in kf.split(x_values):

            x_train = np.array(x_values[train_index])
            y_train = np.array(y_values[train_index])
            x_test = np.array(x_values[test_index])
            y_test = np.array(y_values[test_index])

            model = Sequential()
            model.add(Dense(8, input_shape=(len(features),)))
            model.add(Dense(16, activation=model_name))
            model.add(Dense(32, activation=model_name))
            model.add(Dropout(0.8))
            model.add(Dense(32, activation=model_name))
            model.add(Dense(16, activation=model_name))
            model.add(Dense(8, activation=model_name))
            model.add(Dense(1, activation=model_name))
            model.summary()

            #%%
            model.compile(loss=tf.keras.losses.logcosh, metrics=['accuracy'], optimizer=tf.optimizers.Adam(lr=1e-03))
            model.fit(x_train, y_train, batch_size=16, epochs=200, validation_data=(x_test, y_test),validation_freq=1)

            score = model.evaluate(x_test, y_test, batch_size =16)
            curr_score += score[1]

            if score[1] >= maxaccuracy:
                model.save('./output/model.h5')
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append(P)
                wb.save('./output/K.xlsx')

                maxaccuracy = score[1]

    avg_score = curr_score / w
    print(avg_score)

    return score[1],maxaccuracy

def bian(X,P):

    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x


if __name__ == '__main__':

    path = './data/T/E1.xlsx'
    # path = './data/heart.xlsx'
    workbook = load_workbook(path)
    sheet = workbook.active
    max_column = sheet.max_column
    P=[]
    for i in range(0,max_column-2):
        P.append(1)

    net(P,path)



