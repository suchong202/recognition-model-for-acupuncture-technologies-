import openpyxl
from openpyxl import load_workbook
import numpy as np
import tensorflow as tf
from sklearn.model_selection import train_test_split
from keras import Sequential
from keras.layers import Dense
from keras.layers import Dropout
import filter
import math

#DNN
def net(P:[],path,maxaccuracy):

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values=[]
    y_values=[]
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data)-1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]
    x_values = np.array(x_values[1:])
    y_values = np.array(y_values[1:])
    #x_values = np.array(bian(x_values, P))

    x_train, x_test, y_train, y_test = train_test_split(x_values, y_values, test_size=0.2, random_state=42,
                                                        shuffle=True)
    model_name1 = 'relu'
    model_name2 = 'sigmoid'
    model_name3 = 'tanh'
    model_name4 = 'softmax'
    model_name5 = 'elu'
    model_name6 = 'swish'
    model_name = model_name2
    model = Sequential()
    model.add(Dense(8, input_shape=(len(features),)))
    model.add(Dense(16, activation=model_name))
    model.add(Dense(32, activation=model_name))
    model.add(Dropout(0.8))
    model.add(Dense(32, activation=model_name))
    model.add(Dense(16, activation=model_name))
    model.add(Dense(8, activation=model_name))
    model.add(Dense(1, activation=model_name))
    model.summary()

    model.compile(loss=tf.keras.losses.logcosh, metrics=['accuracy'], optimizer=tf.optimizers.Adam(lr=1e-03))
    history = model.fit(x_train, y_train, batch_size=16, epochs=100, validation_data=(x_test, y_test),validation_freq=1)

    score = model.evaluate(x_test, y_test, batch_size =16)

    print(score[1])

    pre = model.predict(x_values)
    m = 0
    for i in range(0, len(pre)):
        m = m + (pre[i][0] - y_values[i]) ** 2
    loss = math.log(1 + math.exp(-m))
    #print(loss)
    # Using loss or score[1]
    if score[1] >= maxaccuracy:

        model.save('./output/model3.h5')
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(P)
        wb.save('./output/K.xlsx')

        maxaccuracy = score[1]

    return score[1],maxaccuracy

def bian(X,P):

    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x


if __name__ == '__main__':

    path = './data/V/E3.xlsx'
    # path = './data/heart.xlsx'
    workbook = load_workbook(path)
    sheet = workbook.active
    max_column = sheet.max_column
    P=[]
    for i in range(0,max_column-2):
        P.append(1)

    net(P,path,0)

