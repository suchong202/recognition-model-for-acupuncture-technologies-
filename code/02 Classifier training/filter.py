import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pylab as plt
import openpyxl
import numpy as np
from openpyxl import load_workbook
from sklearn.linear_model import Lasso
import scipy.cluster.hierarchy as sch


def min_max_normalize(arr):
    min_val = np.min(arr)
    max_val = np.max(arr)
    normalized_arr = (arr - min_val) / (max_val - min_val)
    return normalized_arr


#读取execl
def getdata(path):
    workbook = load_workbook(path)
    sheet = workbook.active

    D = []
    n = sheet.max_column
    for i in range(0, n):
        A = []
        D.append(A)

    for row in sheet.iter_rows():

        row_data = [cell.value for cell in row]

        for i in range(0, n):

            D[i].append(row_data[i])

    features=[]
    for i in range (1,len(D)):
        features.append(D[i][0])
    data =[]
    for i in range(1, len(D)):
        data.append(min_max_normalize(D[i][1:]))


    return features,data

# Hierarchical Clustering
def Julei(features,data):
    disMat = sch.distance.pdist(data, 'euclidean')
    Z = sch.linkage(disMat, method='average')
    fig = plt.figure()
    P = sch.dendrogram(Z, labels=features)
    print(P)

    plt.show()

#get feature coefficient
def getC(path):

    features,data= getdata(path)
    C=[]
    for i in range(0, len(data)-1):
        distance = np.linalg.norm(np.array(data[i]) - np.array(data[len(data)-1]))
        C.append(distance)
    C=min_max_normalize(C)
    for i in range(0,len(C)):
        C[i]=int(abs(C[i]-0.5)*10)
    C=np.array(C)
    return C

#Julei(features,data)