import openpyxl
import os
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

import Feature
import Read


# RFTR and RDTR or RFLT and RDLT
def model_data1(execlpath,savepath,num):

    W=[]
    channls=8
    for i in range(0,4):
        X=[]
        for i in range(0, channls):
            Y = []
            X.append(Y)
        W.append(X)

    file = os.listdir(execlpath)
    for f in file:

        real_url = execlpath + f
        file1 = os.listdir(real_url)
        for f1 in file1:

            path = real_url + '/' + f1
            D = Read.readexecl(path)

            j = int(f1[0])
            for i in range(0, channls):
                W[j-1][i].append(D[i])


    wb = openpyxl.Workbook()
    sheet = wb.active

    name = ['Num']
    for i in range(1, num + 1):
        name.append(str(i))

    A = ['E-', 'RMS-', 'VAR‘-', 'R‘-', 'FFZ‘-', 'PSV-']
    B = ['P1', 'P2', 'Acc1', 'Acc2', 'Acc3', 'Ang1', 'Ang2', 'Ang3', ]
    C = ['V‘1x', 'V‘1y', 'V‘1z', 'V‘2x', 'V‘2y', 'V‘2z', 'V‘3x', 'V‘3y', 'V‘3z', 'V‘4x', 'V‘4y', 'V‘4z', 'V‘5x', 'V‘5y',
         'V‘5z', 'V‘6x', 'V‘6y', 'V‘6z', 'V‘7x', 'V‘7y', 'V‘7z', 'V‘8x', 'V‘8y', 'V‘8z']
    F1 = ['Num']
    for i in range(0, len(B)):
        for j in range(0, len(A)):
            F1.append(A[j] + B[i])

    F2 = ['Num']
    for i in range(0, len(C)):
        for j in range(0, len(A)):
            F2.append(A[j] + C[i])
    name=F1
    name.append('Type')
    sheet.append(name)

    M = [0, 0, 0, 0]
    n=0

    for z in range(0, 4):
      for b in range(0, 20):

        K = Read.lottery_draw(5, 1, len(W[z][0])-1)
        D = []

        for j in range(0, channls):
            E = []
            for k in range(0, len(K)):
                E.extend(Read.cut(W[z][j][K[k]],0.2))
            D.append(E)
        #print(D)

        L = 50
        for j in range(0, len(D[0])-L):
            #print(j,j+10)
            P1=D[0][j:j+L]
            P2=D[1][j:j+L]

            if Feature.Chu(P1,P2)==1:
                A = []
                A.append(str(n))
                n = n + 1

                for i in range(0, channls):

                    D1 = np.array(D[i][j:j + L])
                    A.append(Feature.E(D1))
                    A.append(Feature.RMS(D1))
                    A.append(Feature.W2(D1))
                    A.append(Feature.W3(D1))
                    A.append(Feature.W7(D1))
                    a, b, c = Feature.findpeaks(D1)
                    A.append((a - c) - (c - b))

                k = z + 1

                if k == 1 or k == 2:
                    A.append(0)
                else:
                    A.append(1)

                M[k - 1] = M[k - 1] + 1
                sheet.append(A)

                print(A)
    print(M)
    wb.save(savepath)



# RFTR or RDTR
def model_data2(execlpath,savepath,num):

    W=[]
    channls=8
    for i in range(0,4):
        X=[]
        for i in range(0, channls):
            Y = []
            X.append(Y)
        W.append(X)

    file = os.listdir(execlpath)
    for f in file:

        real_url = execlpath + f
        file1 = os.listdir(real_url)
        for f1 in file1:

            path = real_url + '/' + f1
            D = Read.readexecl(path)

            j = int(f1[0])
            if j==1 or j==2:
                for i in range(0, len(D)):
                    W[j-1][i].append(D[i])


    wb = openpyxl.Workbook()
    sheet = wb.active

    name = ['Num']
    for i in range(1, num + 1):
        name.append(str(i))

    A = ['E-', 'RMS-', 'VAR‘-', 'R‘-', 'FFZ‘-', 'PSV-']
    B = ['P1', 'P2', 'Acc1', 'Acc2', 'Acc3', 'Ang1', 'Ang2', 'Ang3', ]
    C = ['V‘1x', 'V‘1y', 'V‘1z', 'V‘2x', 'V‘2y', 'V‘2z', 'V‘3x', 'V‘3y', 'V‘3z', 'V‘4x', 'V‘4y', 'V‘4z', 'V‘5x', 'V‘5y',
         'V‘5z', 'V‘6x', 'V‘6y', 'V‘6z', 'V‘7x', 'V‘7y', 'V‘7z', 'V‘8x', 'V‘8y', 'V‘8z']
    F1 = ['Num']
    for i in range(0, len(B)):
        for j in range(0, len(A)):
            F1.append(A[j] + B[i])

    F2 = ['Num']
    for i in range(0, len(C)):
        for j in range(0, len(A)):
            F2.append(A[j] + C[i])
    name = F1
    name.append('Type')
    sheet.append(name)


    n=0

    for z in range(0, 2):
      for b in range(0, 20):

        K = Read.lottery_draw(5, 1, len(W[z][0])-1)
        D = []

        for j in range(0, channls):
            E = []
            for k in range(0, len(K)):
                E.extend(Read.cut(W[z][j][K[k]],0.2))
            D.append(E)
        #print(D)

        L = 50
        for j in range(0, len(D[0]) - L):
                # print(j,j+10)

                P1 = D[0][j:j + L]
                P2 = D[1][j:j + L]

                if Feature.Chu(P1, P2) == 1:
                    A = []
                    A.append(str(n))
                    n = n + 1

                    for i in range(0, channls):
                        D1 = np.array(D[i][j:j + L])
                        A.append(Feature.E(D1))
                        A.append(Feature.RMS(D1))
                        A.append(Feature.W2(D1))
                        A.append(Feature.W3(D1))
                        A.append(Feature.W7(D1))
                        a, b, c = Feature.findpeaks(D1)
                        A.append((a - c) - (c - b))

                    k = z + 1

                    if k == 1 or k == 1:
                        A.append(0)
                    else:
                        A.append(1)

                    sheet.append(A)

                    #print(A)

    wb.save(savepath)


# RFLT or RDLT
def model_data3(execlpath,savepath,num):

    W=[]
    channls=8
    for i in range(0,4):
        X=[]
        for i in range(0, channls):
            Y = []
            X.append(Y)
        W.append(X)

    file = os.listdir(execlpath)
    for f in file:

        real_url = execlpath + f
        file1 = os.listdir(real_url)
        for f1 in file1:

            path = real_url + '/' + f1
            D = Read.readexecl(path)

            j = int(f1[0])
            if j==3 or j==4:
                for i in range(0, len(D)):
                    W[j-1][i].append(D[i])


    wb = openpyxl.Workbook()
    sheet = wb.active

    name = ['Num']
    for i in range(1, num + 1):
        name.append(str(i))

    A = ['E-', 'RMS-', 'VAR‘-', 'R‘-', 'FFZ‘-', 'PSV-']
    B = ['P1', 'P2', 'Acc1', 'Acc2', 'Acc3', 'Ang1', 'Ang2', 'Ang3', ]
    C = ['V‘1x', 'V‘1y', 'V‘1z', 'V‘2x', 'V‘2y', 'V‘2z', 'V‘3x', 'V‘3y', 'V‘3z', 'V‘4x', 'V‘4y', 'V‘4z', 'V‘5x', 'V‘5y',
         'V‘5z', 'V‘6x', 'V‘6y', 'V‘6z', 'V‘7x', 'V‘7y', 'V‘7z', 'V‘8x', 'V‘8y', 'V‘8z']
    F1 = ['Num']
    for i in range(0, len(B)):
        for j in range(0, len(A)):
            F1.append(A[j] + B[i])

    F2 = ['Num']
    for i in range(0, len(C)):
        for j in range(0, len(A)):
            F2.append(A[j] + C[i])
    name = F1
    name.append('Type')
    sheet.append(name)


    n=0

    for z in range(2, 4):
      for b in range(0, 20):

        K = Read.lottery_draw(5, 1, len(W[z][0])-1)
        D = []

        for j in range(0, channls):
            E = []
            for k in range(0, len(K)):
                E.extend(Read.cut(W[z][j][K[k]],0.2))
            D.append(E)
        #print(D)

        L = 50
        for j in range(0, len(D[0]) - L):
                # print(j,j+10)

                P1 = D[0][j:j + L]
                P2 = D[1][j:j + L]

                if Feature.Chu(P1, P2) == 1:
                    A = []
                    A.append(str(n))
                    n = n + 1

                    for i in range(0, channls):
                        D1 = np.array(D[i][j:j + L])
                        A.append(Feature.E(D1))
                        A.append(Feature.RMS(D1))
                        A.append(Feature.W2(D1))
                        A.append(Feature.W3(D1))
                        A.append(Feature.W7(D1))
                        a, b, c = Feature.findpeaks(D1)
                        A.append((a - c) - (c - b))

                    k = z + 1

                    if k == 3 or k == 3:
                        A.append(0)
                    else:
                        A.append(1)

                    sheet.append(A)

                    print(A)

    wb.save(savepath)


if __name__ == '__main__':


    url = './data/T/Expert/'

    save = 'E:/Code/02 Classifier training/data/T/E1.xlsx'
    model_data1(url, save, 8 * 6)

    save = 'E:/Code/02 Classifier training/data/T/E2.xlsx'
    model_data2(url, save, 8 * 6)

    save = 'E:/Code/02 Classifier training/data/T/E3.xlsx'
    model_data3(url, save, 8 * 6)

