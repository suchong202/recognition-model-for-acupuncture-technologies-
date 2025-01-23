import openpyxl
from openpyxl import load_workbook
import numpy as np
import pickle
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.svm import SVC

#SVM
def net(P:[],path,maxaccuracy):

    workbook = load_workbook(path)
    sheet = workbook.active

    x_values = []
    y_values = []
    for row in sheet.iter_rows():
        row_data = [cell.value for cell in row]

        x = row_data[1:-1]
        y = row_data[len(row_data) - 1]

        x_values.append(x)
        y_values.append(y)

    features = x_values[0]
    x_values = np.array(x_values[1:])
    y_values = np.array(y_values[1:])
    x_values = np.array(bian(x_values, P))

    x_train, x_test, y_train, y_test = train_test_split(x_values, y_values, test_size=0.2, random_state=42,shuffle=True)

    model = SVC(kernel='linear', C=1)

    history = model.fit(x_train, y_train)

    y_pred = model.predict(x_test)
    print(y_pred)

    accuracy = accuracy_score(y_test, y_pred)
    print(f"Accuracy: {accuracy:.2f}")

    if accuracy >= maxaccuracy:

        model_path = './output/model.pkl'
        pickle.dump(model, open(model_path, 'wb'))

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(P)
        wb.save('./output/K.xlsx')

        maxaccuracy = accuracy

    return accuracy, maxaccuracy

def bian(X,P):

    new_x =[]
    for i in range(0,len(X)):
        x=[]
        for j in range(0, len(P)):
            x.append(X[i][j]*P[j])
        new_x.append(x)
    return new_x

if __name__ == '__main__':

    path = './data/T/E1.xlsx'
    # path = './data/heart.xlsx'
    workbook = load_workbook(path)
    sheet = workbook.active
    max_column = sheet.max_column
    P = []
    for i in range(0, max_column - 2):
        P.append(1)
    net(P, path, 0)










